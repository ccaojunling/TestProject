import openpyxl


# wb = openpyxl.load_workbook("../data/interface_data.xlsx")
# print(wb)
# print(wb.sheetnames)
# print(wb.active)
# sheet = wb[wb.sheetnames[0]]
# print(sheet.cell(1, 2))


class OperationExcel:
    def __init__(self, filename, index=0):
        self.filename = filename
        self.index = index
        self.data = self.get_data()

    # 获取excel里的数据
    def get_data(self):
        open_excel = openpyxl.load_workbook(self.filename)
        sheet_name = open_excel.sheetnames
        data = open_excel[sheet_name[self.index]]
        return data

    # 获取用例条数
    def get_case_num(self):
        num = self.data.max_row
        return num

    # 获取单元格的内容
    def get_cell_vales(self, row, col):
        result = self.data.cell(row, col).value
        return result

    # 获取某一行的内容
    def get_row_value(self, row):
        row_list = []
        row_data = self.data[row]
        for i in row_data:
            row_list.append(i.value)
        return row_list




if __name__ == '__main__':
    ope = OperationExcel("../data/interface_data.xlsx")
    print(ope.data)
    print(ope.get_cell_vales(1, 2))
    print(ope.get_case_num())
    print(ope.get_row_value(1))


